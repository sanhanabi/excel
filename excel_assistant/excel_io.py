from __future__ import annotations

from collections import Counter
import math
from pathlib import Path
from typing import Any, Iterable
import warnings

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import ColumnProfile, ExecutionResult, TableCandidate, WorkbookProfile
from .output import apply_output_directives


def _is_nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _text_like(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip()) and not value.startswith("=")


def _segments(values: list[Any], minimum_width: int = 2) -> Iterable[tuple[int, int]]:
    start: int | None = None
    for index, value in enumerate(values, start=1):
        if _is_nonempty(value):
            if start is None:
                start = index
        elif start is not None:
            if index - start >= minimum_width:
                yield start, index - 1
            start = None
    if start is not None and len(values) + 1 - start >= minimum_width:
        yield start, len(values)


def _unique_headers(values: list[Any], start_column: int) -> tuple[str, ...]:
    used: dict[str, int] = {}
    result: list[str] = []
    for offset, value in enumerate(values):
        base = str(value).strip() if _is_nonempty(value) else f"열_{start_column + offset}"
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return tuple(result)


def _data_end(ws, start_row: int, start_column: int, end_column: int) -> int:
    last_nonempty = start_row - 1
    blank_run = 0
    for row_number in range(start_row, ws.max_row + 1):
        values = [
            ws.cell(row=row_number, column=column).value
            for column in range(start_column, end_column + 1)
        ]
        if any(_is_nonempty(value) for value in values):
            last_nonempty = row_number
            blank_run = 0
        else:
            blank_run += 1
            if blank_run >= 2:
                break
    return last_nonempty


def _nonempty_rows(ws, start_row: int, end_row: int, start_column: int, end_column: int) -> int:
    return sum(
        any(
            _is_nonempty(ws.cell(row=row_number, column=column).value)
            for column in range(start_column, end_column + 1)
        )
        for row_number in range(start_row, end_row + 1)
    )


def _score_candidate(ws, row: int, start_column: int, end_column: int) -> tuple[float, int]:
    header_values = [
        ws.cell(row=row, column=column).value
        for column in range(start_column, end_column + 1)
    ]
    width = len(header_values)
    text_ratio = sum(_text_like(value) for value in header_values) / width
    unique_ratio = len({str(value).strip() for value in header_values}) / width
    next_rows = []
    for row_number in range(row + 1, min(ws.max_row, row + 3) + 1):
        filled = sum(
            _is_nonempty(ws.cell(row=row_number, column=column).value)
            for column in range(start_column, end_column + 1)
        )
        next_rows.append(filled / width)
    if not next_rows:
        return 0.0, row
    data_density = sum(next_rows) / len(next_rows)
    end_row = _data_end(ws, row + 1, start_column, end_column)
    data_rows = max(0, end_row - row)
    score = text_ratio * 0.45 + unique_ratio * 0.15 + data_density * 0.30
    score += min(data_rows, 10) / 100
    if any(str(value).strip().startswith("[") for value in header_values):
        score -= 0.4
    return max(0.0, min(score, 1.0)), end_row


def detect_tables(path: str | Path) -> list[TableCandidate]:
    """Return one primary data region for each worksheet without modifying the file."""
    workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=False,
        keep_vba=False,
        keep_links=False,
    )
    candidates: list[TableCandidate] = []
    try:
        for ws in workbook.worksheets:
            registered_candidates: list[TableCandidate] = []
            for table in ws.tables.values():
                start_column, header_row, end_column, table_end_row = range_boundaries(
                    table.ref
                )
                data_end_row = table_end_row
                if bool(getattr(table, "totalsRowShown", False)) or bool(
                    getattr(table, "totalsRowCount", 0)
                ):
                    data_end_row -= 1
                header_values = [
                    ws.cell(row=header_row, column=column).value
                    for column in range(start_column, end_column + 1)
                ]
                if data_end_row <= header_row:
                    continue
                registered_candidates.append(
                    TableCandidate(
                        sheet_name=ws.title,
                        header_row=header_row,
                        start_column=start_column,
                        end_column=end_column,
                        data_start_row=header_row + 1,
                        data_end_row=data_end_row,
                        headers=_unique_headers(header_values, start_column),
                        nonempty_row_count=_nonempty_rows(
                            ws,
                            header_row + 1,
                            data_end_row,
                            start_column,
                            end_column,
                        ),
                        confidence=1.0,
                    )
                )
            if registered_candidates:
                primary = max(
                    registered_candidates,
                    key=lambda item: (
                        item.row_count * (item.end_column - item.start_column + 1),
                        item.row_count,
                    ),
                )
                candidates.append(primary)
                continue

            row_limit = min(ws.max_row, 200)
            column_limit = min(ws.max_column, 100)
            sheet_candidates: list[TableCandidate] = []
            for row_number in range(1, row_limit + 1):
                row_values = [
                    ws.cell(row=row_number, column=column).value
                    for column in range(1, column_limit + 1)
                ]
                for start_column, end_column in _segments(row_values):
                    values = row_values[start_column - 1 : end_column]
                    if sum(_text_like(value) for value in values) < 2:
                        continue
                    score, end_row = _score_candidate(
                        ws, row_number, start_column, end_column
                    )
                    if score < 0.62 or end_row <= row_number:
                        continue
                    candidate = TableCandidate(
                        sheet_name=ws.title,
                        header_row=row_number,
                        start_column=start_column,
                        end_column=end_column,
                        data_start_row=row_number + 1,
                        data_end_row=end_row,
                        headers=_unique_headers(values, start_column),
                        nonempty_row_count=_nonempty_rows(
                            ws, row_number + 1, end_row, start_column, end_column
                        ),
                        confidence=round(score, 3),
                    )
                    sheet_candidates.append(candidate)

            if sheet_candidates:
                primary = max(
                    sheet_candidates,
                    key=lambda item: (
                        item.row_count * (item.end_column - item.start_column + 1),
                        item.end_column - item.start_column + 1,
                        -item.header_row,
                        item.confidence,
                    ),
                )
                candidates.append(primary)
    finally:
        workbook.close()
    return candidates


def load_detected_table(
    path: str | Path,
    candidate: TableCandidate,
    *,
    include_source_metadata: bool = False,
    include_hidden_rows: bool = True,
) -> pd.DataFrame:
    workbook = load_workbook(
        filename=path,
        read_only=False,
        data_only=True,
        keep_vba=False,
        keep_links=False,
    )
    try:
        ws = workbook[candidate.sheet_name]
        hidden_rows = {
            int(row_number)
            for row_number, dimension in ws.row_dimensions.items()
            if bool(dimension.hidden)
        }
        rows: list[list[Any]] = []
        source_rows: list[int] = []
        hidden_flags: list[bool] = []
        for row_number in range(candidate.data_start_row, candidate.data_end_row + 1):
            is_hidden = row_number in hidden_rows
            if is_hidden and not include_hidden_rows:
                continue
            values = [
                ws.cell(row=row_number, column=column).value
                for column in range(candidate.start_column, candidate.end_column + 1)
            ]
            if any(_is_nonempty(value) for value in values):
                rows.append(values)
                source_rows.append(row_number)
                hidden_flags.append(is_hidden)
        result = pd.DataFrame(rows, columns=list(candidate.headers)).dropna(how="all")
        if include_source_metadata:
            result["_원본행"] = source_rows
            result["_숨김행"] = hidden_flags
        return result.reset_index(drop=True)
    finally:
        workbook.close()


def combine_detected_tables(
    path: str | Path,
    candidates: list[TableCandidate],
    *,
    include_hidden_rows: bool = True,
) -> pd.DataFrame:
    """Read multiple detected tables and combine them without modifying the workbook."""
    if not candidates:
        raise ValueError("결합할 표가 없습니다.")
    frames: list[pd.DataFrame] = []
    for candidate in candidates:
        frame = load_detected_table(
            path,
            candidate,
            include_source_metadata=True,
            include_hidden_rows=include_hidden_rows,
        )
        frame.insert(0, "원본시트", candidate.sheet_name)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True, sort=False)


def combine_detected_files(
    selections: list[tuple[str | Path, TableCandidate]],
    *,
    include_hidden_rows: bool = True,
) -> pd.DataFrame:
    """Combine selected tables from multiple files while preserving provenance."""
    if not selections:
        raise ValueError("결합할 파일과 표가 없습니다.")
    frames: list[pd.DataFrame] = []
    for raw_path, candidate in selections:
        path = Path(raw_path)
        frame = load_detected_table(
            path,
            candidate,
            include_source_metadata=True,
            include_hidden_rows=include_hidden_rows,
        )
        collisions = {
            "원본파일",
            "원본시트",
        }.intersection(str(column) for column in frame.columns)
        if collisions:
            raise ValueError(
                f"출처 열 이름과 원본 열이 충돌합니다: {', '.join(sorted(collisions))}"
            )
        frame.insert(0, "원본시트", candidate.sheet_name)
        frame.insert(0, "원본파일", path.name)
        frames.append(frame)
    return pd.concat(frames, ignore_index=True, sort=False)


def build_profile(
    df: pd.DataFrame,
    file_name: str,
    sheet_name: str,
    sample_count: int = 3,
) -> WorkbookProfile:
    columns: list[ColumnProfile] = []
    for raw_name in df.columns:
        name = str(raw_name)
        series = df[raw_name]
        nonempty = series.dropna()
        samples: list[Any] = []
        for value in nonempty.drop_duplicates().head(sample_count).tolist():
            samples.append(_profile_scalar(value))
        statistics = _column_statistics(series, nonempty)
        columns.append(
            ColumnProfile(
                name=name,
                dtype=str(series.dtype),
                missing_count=int(series.isna().sum()),
                unique_count=int(series.nunique(dropna=True)),
                sample_values=samples,
                statistics=statistics,
            )
        )
    return WorkbookProfile(
        file_name=file_name,
        sheet_name=sheet_name,
        row_count=len(df),
        columns=columns,
    )


def _profile_scalar(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def _rounded_ratio(value: float) -> float:
    return round(float(value), 4)


def _column_statistics(
    series: pd.Series,
    nonempty: pd.Series,
    top_value_count: int = 3,
) -> dict[str, Any]:
    """Return compact, JSON-safe facts that help a small planner avoid guessing."""
    if nonempty.empty:
        return {}

    if pd.api.types.is_numeric_dtype(series):
        numeric = pd.to_numeric(nonempty, errors="coerce").dropna()
        return {
            "min": _profile_scalar(numeric.min()),
            "max": _profile_scalar(numeric.max()),
            "sum": _profile_scalar(numeric.sum()),
            "mean": _profile_scalar(numeric.mean()),
            "median": _profile_scalar(numeric.median()),
            "zero_count": int(numeric.eq(0).sum()),
            "negative_count": int(numeric.lt(0).sum()),
        }

    if pd.api.types.is_datetime64_any_dtype(series):
        dates = pd.to_datetime(nonempty, errors="coerce").dropna()
        return {
            "min_date": dates.min().isoformat(),
            "max_date": dates.max().isoformat(),
        }

    text = nonempty.astype(str)
    row_count = len(text)
    unique_count = int(text.nunique(dropna=True))
    numeric = pd.to_numeric(text, errors="coerce")
    nonnumeric_text = text[numeric.isna()]
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        parsed_dates = pd.to_datetime(nonnumeric_text, errors="coerce")

    statistics: dict[str, Any] = {
        "unique_ratio": _rounded_ratio(unique_count / row_count),
        "average_length": round(float(text.str.len().mean()), 2),
        "numeric_parse_ratio": _rounded_ratio(numeric.notna().mean()),
        "datetime_parse_ratio": _rounded_ratio(
            parsed_dates.notna().sum() / row_count
        ),
    }
    # Repeated values describe categorical columns well; unique identifiers do not.
    if unique_count <= max(20, row_count // 2):
        statistics["top_values"] = [
            {"value": _profile_scalar(value), "count": int(count)}
            for value, count in text.value_counts(dropna=True)
            .head(top_value_count)
            .items()
        ]
    return statistics


def _source_column_formats(
    source_path: str | Path,
    table: TableCandidate,
) -> dict[str, str]:
    workbook = load_workbook(
        filename=source_path,
        read_only=False,
        data_only=False,
        keep_vba=False,
        keep_links=False,
    )
    try:
        ws = workbook[table.sheet_name]
        formats: dict[str, str] = {}
        for offset, header in enumerate(table.headers):
            column_number = table.start_column + offset
            used_formats = [
                ws.cell(row=row_number, column=column_number).number_format
                for row_number in range(table.data_start_row, table.data_end_row + 1)
                if ws.cell(row=row_number, column=column_number).value is not None
            ]
            if used_formats:
                formats[str(header)] = Counter(used_formats).most_common(1)[0][0]
        return formats
    finally:
        workbook.close()


def _inferred_number_format(series: pd.Series) -> str | None:
    if pd.api.types.is_datetime64_any_dtype(series):
        nonempty = pd.to_datetime(series.dropna())
        has_time = any(
            value.hour or value.minute or value.second or value.microsecond
            for value in nonempty
        )
        return "yyyy-mm-dd hh:mm:ss" if has_time else "yyyy-mm-dd"
    if pd.api.types.is_timedelta64_dtype(series):
        return "[h]:mm:ss"
    return None


def _unique_sheet_name(workbook, requested_name: str) -> str:
    base = requested_name[:31] or "결과"
    if base not in workbook.sheetnames:
        return base
    number = 2
    while True:
        suffix = f"_{number}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate not in workbook.sheetnames:
            return candidate
        number += 1


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if bool(pd.isna(value)):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, pd.Timedelta):
        return value.to_pytimedelta()
    if hasattr(value, "item"):
        return value.item()
    return value


def save_result(
    df: pd.DataFrame | ExecutionResult,
    path: str | Path,
    sheet_name: str = "결과",
    *,
    source_path: str | Path | None = None,
    source_table: TableCandidate | None = None,
    source_paths: list[str | Path] | None = None,
) -> None:
    execution_result = df if isinstance(df, ExecutionResult) else None
    source_df = execution_result.df if execution_result is not None else df
    directives = execution_result.output_directives if execution_result is not None else []
    output_path = Path(path)
    if (
        source_path is not None
        and Path(source_path).resolve() == output_path.resolve()
    ):
        raise ValueError("원본 엑셀 파일에는 결과를 덮어쓸 수 없습니다.")
    if source_paths and any(
        Path(item).resolve() == output_path.resolve() for item in source_paths
    ):
        raise ValueError("병합에 사용한 원본 엑셀 파일에는 결과를 덮어쓸 수 없습니다.")
    output_df = source_df.drop(
        columns=[column for column in source_df.columns if str(column).startswith("_")],
        errors="ignore",
    )
    source_formats = (
        _source_column_formats(source_path, source_table)
        if source_path is not None and source_table is not None
        else {}
    )
    workbook = Workbook()
    ws = workbook.active
    ws.title = sheet_name[:31] or "결과"
    for row_number, row in enumerate(
        dataframe_to_rows(output_df, index=False, header=True),
        start=1,
    ):
        for column_number, value in enumerate(row, start=1):
            ws.cell(
                row=row_number,
                column=column_number,
                value=_excel_value(value),
            )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if ws.max_row > 1 and ws.max_column > 0:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    try:
        for column_number, raw_name in enumerate(output_df.columns, start=1):
            column_name = str(raw_name)
            number_format = source_formats.get(column_name)
            if not number_format or number_format == "General":
                number_format = _inferred_number_format(output_df[raw_name])
            if not number_format:
                continue
            for row_number in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_number, column=column_number)
                if cell.value is not None:
                    cell.number_format = number_format
        apply_output_directives(ws, output_df, directives)
        if any(
            directive.kind in {
                "formula_column",
                "formula_subtotals",
                "formula_date_difference",
                "add_total_row",
                "add_conditional_summary_row",
            }
            for directive in directives
        ):
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
        workbook.save(output_path)
    finally:
        workbook.close()
