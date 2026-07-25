from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .conditions import validate_condition
from .models import OutputDirective, PlanStep
from .operations import _condition_mask


COLORS = {
    "red": "FFFFC7CE",
    "yellow": "FFFFEB9C",
    "green": "FFC6EFCE",
    "blue": "FFBDD7EE",
    "gray": "FFD9E1F2",
}

NUMBER_FORMATS = {
    "currency": "₩#,##0",
    "thousands": "#,##0",
    "percent": "0.00%",
    "percent_points": '0.00"%"',
    "decimal_0": "0",
    "decimal_1": "0.0",
    "decimal_2": "0.00",
    "decimal_3": "0.000",
    "decimal_4": "0.0000",
    "date": "yyyy-mm-dd",
    "datetime": "yyyy-mm-dd hh:mm:ss",
}

COLOR_SCALES = {
    "red_yellow_green": ("FFF8696B", "FFFFEB84", "FF63BE7B"),
    "blue_white": ("FF5B9BD5", "FFFFFFFF", "FFDDEBF7"),
    "green_white": ("FF70AD47", "FFFFFFFF", "FFE2F0D9"),
}

OUTPUT_FUNCTIONS = {
    "highlight_rows",
    "highlight_extremes",
    "highlight_missing",
    "format_numbers",
    "color_scale",
    "add_total_row",
    "add_conditional_summary_row",
}

SUBTOTAL_CODES = {
    "sum": 109,
    "average": 101,
    "count": 103,
    "min": 105,
    "max": 104,
}

CONDITIONAL_AGGREGATES = {
    "sum": "SUMIF",
    "count": "COUNTIF",
    "average": "AVERAGEIF",
}

FORMULA_CONDITION_OPERATORS = {
    "==", "!=", ">", ">=", "<", "<=", "contains", "startswith",
    "endswith", "is_null", "not_null",
}


def _as_columns(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, list) and all(isinstance(item, str) for item in value):
        return value
    return []


def directive_from_step(step: PlanStep) -> OutputDirective:
    if step.function not in OUTPUT_FUNCTIONS:
        raise ValueError(f"출력 지시사항으로 지원하지 않는 함수입니다: {step.function}")
    return OutputDirective(kind=step.function, params=dict(step.params))


def validate_output_step(df: pd.DataFrame, step: PlanStep) -> int | None:
    params = step.params
    if step.function == "highlight_rows":
        validate_condition(params.get("operator"), params.get("value"))
        mask = _condition_mask(
            df[str(params["column"])],
            str(params["operator"]),
            params.get("value"),
        )
        affected_rows = int(mask.sum())
        if affected_rows == 0:
            raise ValueError(
                "강조 조건에 맞는 행이 0개입니다. 조건을 확인해 주세요."
            )
        targets = _as_columns(params.get("target_columns")) or [str(item) for item in df.columns]
        _require_columns(df, targets)
        _require_color(params.get("color"))
        return affected_rows
    elif step.function == "highlight_extremes":
        column = str(params["column"])
        _require_columns(df, [column])
        if params.get("mode") not in {"max", "min", "top_n", "bottom_n"}:
            raise ValueError("극값 강조 방식이 올바르지 않습니다.")
        n = params.get("n", 1)
        if not isinstance(n, int) or n < 1:
            raise ValueError("강조할 개수는 1 이상의 정수여야 합니다.")
        if pd.to_numeric(df[column], errors="coerce").notna().sum() == 0:
            raise ValueError("극값 강조 대상 열에 숫자 값이 없습니다.")
        _require_color(params.get("color"))
    elif step.function == "highlight_missing":
        _require_columns(df, _as_columns(params.get("columns")))
        _require_color(params.get("color"))
    elif step.function == "format_numbers":
        _require_columns(df, _as_columns(params.get("columns")))
        if params.get("format") not in NUMBER_FORMATS:
            raise ValueError("지원하지 않는 숫자 표시 형식입니다.")
    elif step.function == "color_scale":
        column = str(params["column"])
        _require_columns(df, [column])
        if params.get("palette") not in COLOR_SCALES:
            raise ValueError("지원하지 않는 색상 스케일입니다.")
        if pd.to_numeric(df[column], errors="coerce").notna().sum() == 0:
            raise ValueError("색상 스케일 대상 열에 숫자 값이 없습니다.")
    elif step.function == "add_total_row":
        _require_columns(df, _as_columns(params.get("value_columns")))
        if params.get("aggregate", "sum") not in SUBTOTAL_CODES:
            raise ValueError("지원하지 않는 전체 요약 방식입니다.")
    elif step.function == "add_conditional_summary_row":
        output_column = str(params["output_column"])
        aggregate = str(params["aggregate"])
        conditions = _summary_conditions(params)
        required_columns = [
            output_column,
            *(str(condition["column"]) for condition in conditions),
        ]
        value_column = params.get("value_column")
        if aggregate in {"sum", "average"}:
            if not isinstance(value_column, str) or not value_column:
                raise ValueError("조건부 합계·평균에는 값 열이 필요합니다.")
            required_columns.append(value_column)
        if params.get("label_column") is not None:
            required_columns.append(str(params["label_column"]))
        _require_columns(df, required_columns)
        masks: list[pd.Series] = []
        for condition in conditions:
            validate_condition(condition.get("operator"), condition.get("value"))
            if condition.get("operator") not in FORMULA_CONDITION_OPERATORS:
                raise ValueError("조건부 요약 수식에서 지원하지 않는 비교 방식입니다.")
            masks.append(
                _condition_mask(
                    df[str(condition["column"])],
                    str(condition["operator"]),
                    condition.get("value"),
                )
            )
        if aggregate not in CONDITIONAL_AGGREGATES:
            raise ValueError("조건부 요약은 sum, count, average만 지원합니다.")
        combined = masks[0]
        for mask in masks[1:]:
            combined = combined & mask
        return int(combined.sum())
    else:
        raise ValueError(f"지원하지 않는 출력 함수입니다: {step.function}")
    return None


def _require_columns(df: pd.DataFrame, columns: list[str]) -> None:
    if not columns:
        raise ValueError("대상 열이 하나 이상 필요합니다.")
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"출력 대상 열을 찾을 수 없습니다: {', '.join(missing)}")
    internal = [column for column in columns if column.startswith("_")]
    if internal:
        raise ValueError(
            f"저장 전에 제거되는 내부 열에는 서식을 적용할 수 없습니다: "
            f"{', '.join(internal)}"
        )


def _require_color(color: Any) -> None:
    if color not in COLORS:
        raise ValueError("지원하지 않는 강조 색상입니다.")


def apply_output_directives(
    ws,
    df: pd.DataFrame,
    directives: list[OutputDirective],
) -> None:
    headers = {str(value): index for index, value in enumerate(df.columns, start=1)}
    for directive in directives:
        if directive.kind == "formula_column":
            _apply_formula_column(ws, headers, directive.params)
        elif directive.kind == "formula_subtotals":
            _apply_formula_subtotals(ws, df, headers, directive.params)
        elif directive.kind == "formula_date_difference":
            _apply_formula_date_difference(ws, headers, directive.params)
        elif directive.kind == "add_total_row":
            _apply_total_row(ws, df, headers, directive.params)
        elif directive.kind == "add_conditional_summary_row":
            _apply_conditional_summary_row(ws, df, headers, directive.params)

    for directive in directives:
        if directive.kind == "highlight_rows":
            _apply_highlight_rows(ws, df, headers, directive.params)
        elif directive.kind == "highlight_extremes":
            _apply_highlight_extremes(ws, df, headers, directive.params)
        elif directive.kind == "highlight_missing":
            _apply_highlight_missing(ws, df, headers, directive.params)
        elif directive.kind == "format_numbers":
            _apply_number_format(ws, headers, directive.params)
        elif directive.kind == "color_scale":
            _apply_color_scale(ws, headers, directive.params)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=COLORS[color])


def _apply_highlight_rows(ws, df, headers, params) -> None:
    mask = _condition_mask(
        df[str(params["column"])], str(params["operator"]), params.get("value")
    )
    targets = _as_columns(params.get("target_columns")) or list(headers)
    fill = _fill(str(params["color"]))
    for excel_row, matched in enumerate(mask.tolist(), start=2):
        if not matched:
            continue
        for column in targets:
            ws.cell(excel_row, headers[column]).fill = fill


def _extreme_indices(series: pd.Series, mode: str, n: int, include_ties: bool) -> set[Any]:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    if numeric.empty:
        return set()
    if mode == "max":
        return set(numeric.index[numeric.eq(numeric.max())].tolist())
    if mode == "min":
        return set(numeric.index[numeric.eq(numeric.min())].tolist())
    largest = mode == "top_n"
    ordered = numeric.nlargest(n) if largest else numeric.nsmallest(n)
    if not include_ties or len(ordered) < n:
        return set(ordered.index.tolist())
    threshold = ordered.iloc[-1]
    mask = numeric.ge(threshold) if largest else numeric.le(threshold)
    return set(numeric.index[mask].tolist())


def _apply_highlight_extremes(ws, df, headers, params) -> None:
    column = str(params["column"])
    indices = _extreme_indices(
        df[column],
        str(params["mode"]),
        int(params.get("n", 1)),
        bool(params.get("include_ties", True)),
    )
    fill = _fill(str(params["color"]))
    for excel_row, data_index in enumerate(df.index, start=2):
        if data_index in indices:
            ws.cell(excel_row, headers[column]).fill = fill


def _apply_highlight_missing(ws, df, headers, params) -> None:
    fill = _fill(str(params["color"]))
    for column in _as_columns(params["columns"]):
        missing = df[column].isna() | df[column].astype("string").str.strip().eq("")
        for excel_row, is_missing in enumerate(missing.tolist(), start=2):
            if is_missing:
                ws.cell(excel_row, headers[column]).fill = fill


def _apply_number_format(ws, headers, params) -> None:
    number_format = NUMBER_FORMATS[str(params["format"])]
    for column in _as_columns(params["columns"]):
        column_number = headers[column]
        for row_number in range(2, ws.max_row + 1):
            if ws.cell(row_number, column_number).value is not None:
                ws.cell(row_number, column_number).number_format = number_format


def _apply_color_scale(ws, headers, params) -> None:
    column_number = headers[str(params["column"])]
    column_letter = get_column_letter(column_number)
    start, middle, end = COLOR_SCALES[str(params["palette"])]
    ws.conditional_formatting.add(
        f"{column_letter}2:{column_letter}{ws.max_row}",
        ColorScaleRule(
            start_type="min",
            start_color=start,
            mid_type="percentile",
            mid_value=50,
            mid_color=middle,
            end_type="max",
            end_color=end,
        ),
    )


def _apply_total_row(ws, df, headers, params) -> None:
    row_number = ws.max_row + 1
    data_end_row = len(df) + 1
    aggregate = str(params.get("aggregate", "sum"))
    subtotal_code = SUBTOTAL_CODES[aggregate]
    ws.cell(row_number, 1, str(params.get("label", "전체 합계")))
    ws.cell(row_number, 1).font = Font(bold=True)
    for column in _as_columns(params["value_columns"]):
        column_number = headers[column]
        letter = get_column_letter(column_number)
        ws.cell(
            row_number,
            column_number,
            f"=SUBTOTAL({subtotal_code},{letter}2:{letter}{data_end_row})",
        )
        ws.cell(row_number, column_number).font = Font(bold=True)


def _formula_criteria(operator: str, value: Any) -> str:
    def quoted(text: str) -> str:
        return '"' + text.replace('"', '""') + '"'

    def escaped_text(item: Any) -> str:
        return str(item).replace("~", "~~").replace("*", "~*").replace("?", "~?")

    if operator == "is_null":
        return '""'
    if operator == "not_null":
        return '"<>"'
    if operator == "contains":
        return quoted(f"*{escaped_text(value)}*")
    if operator == "startswith":
        return quoted(f"{escaped_text(value)}*")
    if operator == "endswith":
        return quoted(f"*{escaped_text(value)}")
    if operator == "==":
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            return str(value)
        return quoted(str(value))
    symbol = "<>" if operator == "!=" else operator
    return quoted(f"{symbol}{value}")


def _summary_conditions(params: dict[str, Any]) -> list[dict[str, Any]]:
    raw_conditions = params.get("conditions")
    legacy_present = any(
        key in params for key in ("condition_column", "operator", "value")
    )
    if raw_conditions is not None and legacy_present:
        raise ValueError("단일 조건과 다중 조건을 동시에 사용할 수 없습니다.")
    if raw_conditions is not None:
        if not isinstance(raw_conditions, list) or not raw_conditions:
            raise ValueError("조건부 요약 조건이 하나 이상 필요합니다.")
        if params.get("logic", "and") != "and":
            raise ValueError("조건부 요약의 다중 조건은 AND만 지원합니다.")
        conditions = raw_conditions
    else:
        if not all(key in params for key in ("condition_column", "operator", "value")):
            raise ValueError("조건부 요약의 단일 조건이 완전하지 않습니다.")
        conditions = [
            {
                "column": params["condition_column"],
                "operator": params["operator"],
                "value": params.get("value"),
            }
        ]
    if not all(isinstance(condition, dict) for condition in conditions):
        raise ValueError("조건부 요약 조건 형식이 올바르지 않습니다.")
    return conditions


def _apply_conditional_summary_row(ws, df, headers, params) -> None:
    row_number = ws.max_row + 1
    data_end_row = len(df) + 1
    output_column = str(params["output_column"])
    output_number = headers[output_column]
    label_column = params.get("label_column")
    if label_column is None:
        label_number = next(
            (number for number in headers.values() if number != output_number),
            None,
        )
    else:
        label_number = headers[str(label_column)]
    if label_number is not None:
        ws.cell(row_number, label_number, str(params.get("label", "조건부 요약")))
        ws.cell(row_number, label_number).font = Font(bold=True)

    conditions = _summary_conditions(params)
    formula_conditions: list[tuple[str, str]] = []
    for condition in conditions:
        condition_number = headers[str(condition["column"])]
        condition_letter = get_column_letter(condition_number)
        condition_range = f"{condition_letter}2:{condition_letter}{data_end_row}"
        criteria = _formula_criteria(
            str(condition["operator"]), condition.get("value")
        )
        formula_conditions.append((condition_range, criteria))
    aggregate = str(params["aggregate"])
    if len(formula_conditions) == 1:
        condition_range, criteria = formula_conditions[0]
        if aggregate == "count":
            formula = f"=COUNTIF({condition_range},{criteria})"
        else:
            value_number = headers[str(params["value_column"])]
            value_letter = get_column_letter(value_number)
            value_range = f"{value_letter}2:{value_letter}{data_end_row}"
            function_name = CONDITIONAL_AGGREGATES[aggregate]
            formula = f"={function_name}({condition_range},{criteria},{value_range})"
    else:
        condition_arguments = ",".join(
            item for pair in formula_conditions for item in pair
        )
        if aggregate == "count":
            formula = f"=COUNTIFS({condition_arguments})"
        else:
            value_number = headers[str(params["value_column"])]
            value_letter = get_column_letter(value_number)
            value_range = f"{value_letter}2:{value_letter}{data_end_row}"
            function_name = "SUMIFS" if aggregate == "sum" else "AVERAGEIFS"
            formula = f"={function_name}({value_range},{condition_arguments})"
    ws.cell(row_number, output_number, formula)
    ws.cell(row_number, output_number).font = Font(bold=True)


def _apply_formula_column(ws, headers, params) -> None:
    result_number = headers[str(params["result_column"])]
    left_number = headers[str(params["left_column"])]
    right_column = params.get("right_column")
    operator = str(params["operator"])
    symbols = {"add": "+", "subtract": "-", "multiply": "*", "divide": "/"}
    for row_number in range(2, ws.max_row + 1):
        left = f"{get_column_letter(left_number)}{row_number}"
        if right_column is not None:
            right = f"{get_column_letter(headers[str(right_column)])}{row_number}"
        else:
            right = str(params["value"])
        if operator == "absolute_difference":
            formula = f"=ABS({left}-{right})"
        elif operator == "percent_of":
            formula = f'=IFERROR({left}/{right}*100,"")'
        elif operator == "divide":
            formula = f'=IFERROR({left}/{right},"")'
        else:
            formula = f"={left}{symbols[operator]}{right}"
        ws.cell(row_number, result_number, formula)


def _apply_formula_date_difference(ws, headers, params) -> None:
    result_number = headers[str(params["result_column"])]
    start_number = headers[str(params["start_column"])]
    end_mode = str(params.get("end_mode", "column"))
    unit = str(params.get("unit", "days"))
    absolute = bool(params.get("absolute", False))
    for row_number in range(2, ws.max_row + 1):
        start = f"{get_column_letter(start_number)}{row_number}"
        if end_mode == "today":
            end = "TODAY()"
            blank_guard = f'{start}=""'
        else:
            end_number = headers[str(params["end_column"])]
            end = f"{get_column_letter(end_number)}{row_number}"
            blank_guard = f'OR({start}="",{end}="")'
        if unit == "months":
            if absolute:
                calculation = f'DATEDIF(MIN({start},{end}),MAX({start},{end}),"m")'
            else:
                calculation = (
                    f'IF({end}>={start},DATEDIF({start},{end},"m"),'
                    f'-DATEDIF({end},{start},"m"))'
                )
        else:
            calculation = f"{end}-{start}"
            if absolute:
                calculation = f"ABS({calculation})"
            if unit == "weeks":
                calculation = f"({calculation})/7"
        ws.cell(
            row_number,
            result_number,
            f'=IF({blank_guard},"",{calculation})',
        )


def _apply_formula_subtotals(ws, df, headers, params) -> None:
    groups = _as_columns(params["group_columns"])
    values = _as_columns(params["value_columns"])
    label = str(params.get("label", "소계"))
    first_group = groups[0]
    group_start = 2
    for data_index, value in enumerate(df[first_group].tolist(), start=2):
        if value != label and value != "전체 합계":
            continue
        range_start = 2 if value == "전체 합계" else group_start
        for column in values:
            column_number = headers[column]
            letter = get_column_letter(column_number)
            ws.cell(
                data_index,
                column_number,
                f"=SUBTOTAL(109,{letter}{range_start}:{letter}{data_index - 1})",
            )
            ws.cell(data_index, column_number).font = Font(bold=True)
        ws.cell(data_index, headers[first_group]).font = Font(bold=True)
        if value == label:
            group_start = data_index + 1
