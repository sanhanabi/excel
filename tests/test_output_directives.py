import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from excel_assistant.excel_io import save_result
from excel_assistant.executor import execute_plan
from excel_assistant.models import ExecutionPlan, PlanStep
from excel_assistant.presentation import format_plan
from excel_assistant.validation import (
    PlanValidationError,
    validate_plan,
    validate_plan_against_data,
)


class OutputDirectiveTests(unittest.TestCase):
    def test_highlight_rows_rejects_null_equality(self):
        frame = pd.DataFrame({"상태": ["완료", None]})
        plan = ExecutionPlan(
            goal="빈 상태 강조",
            steps=[
                PlanStep(
                    "highlight_rows",
                    {
                        "column": "상태",
                        "operator": "==",
                        "value": None,
                        "color": "red",
                    },
                )
            ],
        )

        with self.assertRaisesRegex(PlanValidationError, "is_null 또는 not_null"):
            validate_plan_against_data(frame, plan)

    def test_highlight_rows_rejects_zero_matches(self):
        frame = pd.DataFrame({"상태": ["완료", "완료"]})
        plan = ExecutionPlan(
            goal="미납 강조",
            steps=[
                PlanStep(
                    "highlight_rows",
                    {
                        "column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "color": "red",
                    },
                )
            ],
        )

        with self.assertRaisesRegex(PlanValidationError, "강조 조건에 맞는 행이 0개"):
            validate_plan_against_data(frame, plan)

    def test_highlight_preview_reports_affected_rows(self):
        frame = pd.DataFrame({"상태": ["완료", "미납", "미납"]})
        plan = ExecutionPlan(
            goal="미납 강조",
            steps=[
                PlanStep(
                    "highlight_rows",
                    {
                        "column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "color": "red",
                    },
                )
            ],
        )

        preview = validate_plan_against_data(frame, plan)

        self.assertEqual(preview.steps[0].affected_rows, 2)
        self.assertIn("강조 대상: 2행", format_plan(plan, preview))

    def test_highlights_formats_color_scale_and_total_formula(self):
        frame = pd.DataFrame(
            {
                "상태": ["완납", "미납", "완납"],
                "금액": [1000, 2500, 1500],
                "비고": ["정상", None, "정상"],
            },
            index=[10, 20, 30],
        )
        plan = ExecutionPlan(
            goal="미납과 빈칸을 강조하고 금액을 표시",
            steps=[
                PlanStep(
                    "highlight_rows",
                    {
                        "column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "color": "red",
                    },
                ),
                PlanStep(
                    "highlight_missing",
                    {"columns": "비고", "color": "yellow"},
                ),
                PlanStep(
                    "format_numbers",
                    {"columns": "금액", "format": "thousands"},
                ),
                PlanStep(
                    "color_scale",
                    {"column": "금액", "palette": "red_yellow_green"},
                ),
                PlanStep(
                    "add_total_row",
                    {"value_columns": "금액", "label": "전체 합계"},
                ),
            ],
        )
        validate_plan_against_data(frame, plan)
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "styled.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(ws["A3"].fill.fgColor.rgb, "FFFFC7CE")
                self.assertEqual(ws["C3"].fill.fgColor.rgb, "FFFFEB9C")
                self.assertEqual(ws["B2"].number_format, "#,##0")
                self.assertEqual(len(ws.conditional_formatting), 1)
                self.assertEqual(ws["A5"].value, "전체 합계")
                self.assertEqual(ws["B5"].value, "=SUBTOTAL(109,B2:B4)")
            finally:
                workbook.close()

    def test_calculated_column_keeps_static_preview_and_writes_safe_formulas(self):
        frame = pd.DataFrame({"단가": [100, 200], "수량": [2, 3]})
        plan = ExecutionPlan(
            goal="총액 계산식을 추가",
            steps=[
                PlanStep(
                    "calculate_column",
                    {
                        "result_column": "총액",
                        "operator": "multiply",
                        "left_column": "단가",
                        "right_column": "수량",
                        "value": None,
                        "as_formula": True,
                    },
                )
            ],
        )
        preview = validate_plan_against_data(frame, plan)
        self.assertEqual(preview.final_rows, 2)
        result = execute_plan(frame, plan)
        self.assertEqual(result.df["총액"].tolist(), [200, 600])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "formula.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(ws["C2"].value, "=A2*B2")
                self.assertEqual(ws["C3"].value, "=A3*B3")
                self.assertTrue(workbook.calculation.fullCalcOnLoad)
                self.assertTrue(workbook.calculation.forceFullCalc)
            finally:
                workbook.close()

    def test_total_row_supports_all_subtotal_aggregates_without_chaining_ranges(self):
        frame = pd.DataFrame({"항목": ["A", "B", "C"], "금액": [10, 20, 30]})
        aggregates = [
            ("sum", 109),
            ("average", 101),
            ("count", 103),
            ("min", 105),
            ("max", 104),
        ]
        plan = ExecutionPlan(
            goal="여러 전체 요약행 추가",
            steps=[
                PlanStep(
                    "add_total_row",
                    {
                        "value_columns": "금액",
                        "aggregate": aggregate,
                        "label": aggregate,
                    },
                )
                for aggregate, _ in aggregates
            ],
        )
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "aggregates.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                for row_number, (_, subtotal_code) in enumerate(aggregates, start=5):
                    self.assertEqual(
                        ws.cell(row_number, 2).value,
                        f"=SUBTOTAL({subtotal_code},B2:B4)",
                    )
            finally:
                workbook.close()

    def test_conditional_summary_rows_write_sumif_countif_and_averageif(self):
        frame = pd.DataFrame(
            {
                "상태": ["미납", "완납", "미납"],
                "금액": [100, 200, 300],
                "비고": [None, None, None],
            }
        )
        plan = ExecutionPlan(
            goal="미납 조건부 요약",
            steps=[
                PlanStep(
                    "add_conditional_summary_row",
                    {
                        "condition_column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "aggregate": "sum",
                        "value_column": "금액",
                        "output_column": "금액",
                        "label": "미납 합계",
                    },
                ),
                PlanStep(
                    "add_conditional_summary_row",
                    {
                        "condition_column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "aggregate": "count",
                        "output_column": "금액",
                        "label": "미납 건수",
                    },
                ),
                PlanStep(
                    "add_conditional_summary_row",
                    {
                        "condition_column": "상태",
                        "operator": "==",
                        "value": "미납",
                        "aggregate": "average",
                        "value_column": "금액",
                        "output_column": "금액",
                        "label": "미납 평균",
                    },
                ),
            ],
        )

        preview = validate_plan_against_data(frame, plan)
        self.assertEqual([item.affected_rows for item in preview.steps], [2, 2, 2])
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "conditional.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(ws["B5"].value, '=SUMIF(A2:A4,"미납",B2:B4)')
                self.assertEqual(ws["B6"].value, '=COUNTIF(A2:A4,"미납")')
                self.assertEqual(ws["B7"].value, '=AVERAGEIF(A2:A4,"미납",B2:B4)')
                self.assertEqual(ws["A5"].value, "미납 합계")
                self.assertTrue(workbook.calculation.fullCalcOnLoad)
            finally:
                workbook.close()

    def test_multi_condition_summary_writes_ifs_formulas(self):
        frame = pd.DataFrame(
            {
                "상태": ["미납", "미납", "완납", "미납"],
                "지역": ["서울", "부산", "서울", "서울"],
                "금액": [100, 200, 300, 400],
            }
        )
        conditions = [
            {"column": "상태", "operator": "==", "value": "미납"},
            {"column": "지역", "operator": "==", "value": "서울"},
        ]
        plan = ExecutionPlan(
            goal="서울 미납 다중 조건 요약",
            steps=[
                PlanStep(
                    "add_conditional_summary_row",
                    {
                        "conditions": conditions,
                        "logic": "and",
                        "aggregate": aggregate,
                        "value_column": "금액" if aggregate != "count" else None,
                        "output_column": "금액",
                        "label": label,
                    },
                )
                for aggregate, label in (
                    ("sum", "서울 미납 합계"),
                    ("count", "서울 미납 건수"),
                    ("average", "서울 미납 평균"),
                )
            ],
        )

        preview = validate_plan_against_data(frame, plan)
        self.assertEqual([item.affected_rows for item in preview.steps], [2, 2, 2])
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "multi_conditions.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(
                    ws["C6"].value,
                    '=SUMIFS(C2:C5,A2:A5,"미납",B2:B5,"서울")',
                )
                self.assertEqual(
                    ws["C7"].value,
                    '=COUNTIFS(A2:A5,"미납",B2:B5,"서울")',
                )
                self.assertEqual(
                    ws["C8"].value,
                    '=AVERAGEIFS(C2:C5,A2:A5,"미납",B2:B5,"서울")',
                )
            finally:
                workbook.close()

    def test_today_date_difference_is_saved_as_live_excel_formula(self):
        frame = pd.DataFrame({"접수일": ["2026-01-01", None]})
        plan = ExecutionPlan(
            goal="접수 후 경과일 계산",
            steps=[
                PlanStep(
                    "calculate_date_difference",
                    {
                        "start_column": "접수일",
                        "result_column": "경과일",
                        "end_mode": "today",
                        "unit": "days",
                        "absolute": False,
                        "as_formula": True,
                    },
                )
            ],
        )

        preview = validate_plan_against_data(frame, plan)
        self.assertEqual(preview.final_rows, 2)
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "today.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(ws["B2"].value, '=IF(A2="","",TODAY()-A2)')
                self.assertEqual(ws["B3"].value, '=IF(A3="","",TODAY()-A3)')
                self.assertTrue(workbook.calculation.fullCalcOnLoad)
            finally:
                workbook.close()

    def test_subtotals_can_be_exported_as_subtotal_formulas(self):
        frame = pd.DataFrame({"그룹": ["A", "A", "B"], "금액": [10, 20, 5]})
        plan = ExecutionPlan(
            goal="그룹 소계를 수식으로 추가",
            steps=[
                PlanStep(
                    "add_subtotals",
                    {
                        "group_columns": "그룹",
                        "value_columns": "금액",
                        "label": "소계",
                        "include_grand_total": True,
                        "as_formula": True,
                    },
                )
            ],
        )
        result = execute_plan(frame, plan)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "subtotals.xlsx"
            save_result(result, path)
            workbook = load_workbook(path, data_only=False)
            try:
                ws = workbook.active
                self.assertEqual(ws["B4"].value, "=SUBTOTAL(109,B2:B3)")
                self.assertEqual(ws["B6"].value, "=SUBTOTAL(109,B5:B5)")
                self.assertEqual(ws["B7"].value, "=SUBTOTAL(109,B2:B6)")
            finally:
                workbook.close()

    def test_data_step_after_output_step_is_rejected(self):
        plan = ExecutionPlan(
            goal="잘못된 단계 순서",
            steps=[
                PlanStep(
                    "format_numbers",
                    {"columns": "금액", "format": "thousands"},
                ),
                PlanStep("sort_rows", {"columns": "금액"}),
            ],
        )
        with self.assertRaisesRegex(PlanValidationError, "먼저 와야 합니다"):
            validate_plan(plan, ["금액"])

    def test_output_format_cannot_target_internal_metadata_column(self):
        frame = pd.DataFrame({"금액": [1], "_원본행": [2]})
        plan = ExecutionPlan(
            goal="내부 열 서식 금지",
            steps=[
                PlanStep(
                    "format_numbers",
                    {"columns": "_원본행", "format": "thousands"},
                )
            ],
        )
        with self.assertRaisesRegex(PlanValidationError, "내부 열"):
            validate_plan_against_data(frame, plan)


if __name__ == "__main__":
    unittest.main()
