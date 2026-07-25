import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_assistant.excel_io import (
    combine_detected_files,
    combine_detected_tables,
    detect_tables,
    save_result,
)
from excel_assistant.app import (
    ExcelAssistantApp,
    apply_column_aliases,
    build_cross_sheet_column_aliases,
    infer_cross_sheet_column_aliases,
)
from excel_assistant.models import TableCandidate


class MultiSheetTests(unittest.TestCase):
    @staticmethod
    def _candidate(sheet_name, headers):
        return TableCandidate(
            sheet_name=sheet_name,
            header_row=1,
            start_column=1,
            end_column=len(headers),
            data_start_row=2,
            data_end_row=10,
            headers=tuple(headers),
            nonempty_row_count=9,
            confidence=1.0,
        )

    def test_infers_renamed_column_at_same_position_across_parallel_sheets(self):
        selected = self._candidate(
            "DEC23",
            ["#", "Date rcvd", "Type", "CO", "Amount"],
        )
        parallel = self._candidate(
            "FEB",
            ["#", "Date rcvd", "Type", "COMPANY", "Amount"],
        )
        request = "Date rcvd와 COMPANY가 모두 존재하는 행만 유지"

        aliases = infer_cross_sheet_column_aliases(
            request,
            selected,
            [selected, parallel],
        )

        self.assertEqual(aliases, {"COMPANY": "CO"})
        self.assertEqual(
            apply_column_aliases(request, aliases),
            "Date rcvd와 CO가 모두 존재하는 행만 유지",
        )
        self.assertEqual(
            build_cross_sheet_column_aliases(selected, [selected, parallel]),
            {"CO": ["COMPANY"]},
        )

    def test_ignores_same_width_sheet_when_most_headers_do_not_match(self):
        selected = self._candidate(
            "DEC23",
            ["#", "Date rcvd", "Type", "CO", "Amount"],
        )
        unrelated = self._candidate(
            "요약",
            ["순번", "월", "부서", "COMPANY", "합계"],
        )

        self.assertEqual(
            build_cross_sheet_column_aliases(selected, [selected, unrelated]),
            {},
        )

    def test_does_not_infer_alias_from_different_width_schema(self):
        selected = self._candidate("DEC23", ["Date rcvd", "CO", "Amount"])
        different_width = self._candidate(
            "MAY",
            ["Date rcvd", "COMPANY", "Due Date", "Amount"],
        )

        aliases = infer_cross_sheet_column_aliases(
            "COMPANY와 Due Date가 있는 행만 유지",
            selected,
            [selected, different_width],
        )

        self.assertEqual(aliases, {})

    def test_korean_month_range_selects_matching_english_sheet_names(self):
        candidates = [
            TableCandidate(
                sheet_name=name,
                header_row=1,
                start_column=1,
                end_column=2,
                data_start_row=2,
                data_end_row=3,
                headers=("회사", "금액"),
                nonempty_row_count=2,
                confidence=1.0,
            )
            for name in ("DEC23", "JAN", "FEB", "MAR", "APR", "MAY", "JUNE")
        ]

        class FakeApp:
            _table_candidates = candidates

        selected = ExcelAssistantApp._candidates_for_request(
            FakeApp(), "1월부터 6월까지 시트를 합쳐줘"
        )
        self.assertEqual(
            [candidate.sheet_name for candidate in selected],
            ["JAN", "FEB", "MAR", "APR", "MAY", "JUNE"],
        )

    def test_combines_detected_tables_and_preserves_hidden_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "ledger.xlsx"
            workbook = Workbook()
            for index, sheet_name in enumerate(("1월", "2월")):
                ws = workbook.active if index == 0 else workbook.create_sheet()
                ws.title = sheet_name
                ws.append(["거래일", "회사", "금액"])
                ws.append([f"2026-0{index + 1}-01", "A", 100 + index])
                ws.append([f"2026-0{index + 1}-02", "B", 200 + index])
                ws.append([f"2026-0{index + 1}-03", "C", 300 + index])
            workbook["1월"].row_dimensions[3].hidden = True
            workbook.save(source_path)
            source_bytes_before = source_path.read_bytes()

            candidates = detect_tables(source_path)
            combined = combine_detected_tables(source_path, candidates)

            self.assertEqual(set(combined["원본시트"]), {"1월", "2월"})
            self.assertEqual(len(combined), 6)
            self.assertEqual(int(combined["_숨김행"].sum()), 1)
            self.assertEqual(source_path.read_bytes(), source_bytes_before)

    def test_combines_multiple_files_with_file_and_sheet_provenance(self):
        with tempfile.TemporaryDirectory() as directory:
            selections = []
            source_bytes = {}
            for file_number in (1, 2):
                source_path = Path(directory) / f"ledger_{file_number}.xlsx"
                workbook = Workbook()
                ws = workbook.active
                ws.title = f"{file_number}월"
                ws.append(["회사", "금액"])
                ws.append(["A", file_number * 100])
                workbook.save(source_path)
                source_bytes[source_path] = source_path.read_bytes()
                selections.append((source_path, detect_tables(source_path)[0]))

            combined = combine_detected_files(selections)

            self.assertEqual(
                combined[["원본파일", "원본시트"]].to_dict(orient="records"),
                [
                    {"원본파일": "ledger_1.xlsx", "원본시트": "1월"},
                    {"원본파일": "ledger_2.xlsx", "원본시트": "2월"},
                ],
            )
            self.assertEqual(combined["금액"].tolist(), [100, 200])
            for path, before in source_bytes.items():
                self.assertEqual(path.read_bytes(), before)

    def test_combined_result_cannot_overwrite_any_source_file(self):
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "first.xlsx"
            second = Path(directory) / "second.xlsx"
            Workbook().save(first)
            Workbook().save(second)
            before = second.read_bytes()

            with self.assertRaisesRegex(ValueError, "병합에 사용한 원본"):
                save_result(
                    pd.DataFrame({"금액": [1]}),
                    second,
                    source_paths=[first, second],
                )

            self.assertEqual(second.read_bytes(), before)

    def test_result_file_omits_internal_source_metadata(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "source.xlsx"
            result_path = Path(directory) / "result.xlsx"
            workbook = Workbook()
            workbook.save(source_path)
            frame = pd.DataFrame(
                {
                    "원본시트": ["1월"],
                    "금액": [100],
                    "_원본행": [2],
                    "_숨김행": [False],
                }
            )
            save_result(frame, result_path, source_path=source_path)
            saved = load_workbook(result_path, read_only=True)
            try:
                headers = [cell.value for cell in next(saved.active.iter_rows())]
                self.assertEqual(headers, ["원본시트", "금액"])
            finally:
                saved.close()


if __name__ == "__main__":
    unittest.main()
