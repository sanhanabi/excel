import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from excel_assistant.excel_io import detect_tables, load_detected_table


class TableDetectionTests(unittest.TestCase):
    def test_registered_excel_table_is_primary_and_sheet_appears_once(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registered.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "장부"
            sheet.append(["날짜", "회사", "금액"])
            sheet.append(["2026-01-01", "A", 100])
            sheet.append(["2026-01-02", "B", 200])
            sheet.append(["2026-01-03", "C", 300])
            sheet.append([])
            sheet.append(["HOLD", "실제 데이터처럼 보이는 가짜 제목"])
            sheet.append(["VALUE", "다음 행"])
            sheet.add_table(Table(displayName="LedgerTable", ref="A1:C4"))
            workbook.save(path)

            candidates = detect_tables(path)

            self.assertEqual(len(candidates), 1)
            self.assertEqual(candidates[0].sheet_name, "장부")
            self.assertEqual(candidates[0].header_row, 1)
            self.assertEqual(candidates[0].data_end_row, 4)
            self.assertEqual(candidates[0].headers, ("날짜", "회사", "금액"))

    def test_detects_title_row_and_real_header(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "판매"
            sheet["A1"] = "[표1]"
            sheet.append(["거래처명", "결제금액", "거래일"])
            sheet.append(["가상사", 100000, "2026-01-01"])
            sheet.append(["나유통", 200000, "2026-01-02"])
            workbook.create_sheet("빈 시트")
            workbook.save(path)

            candidates = detect_tables(path)

            self.assertEqual(len(candidates), 1)
            candidate = candidates[0]
            self.assertEqual(candidate.sheet_name, "판매")
            self.assertEqual(candidate.header_row, 2)
            self.assertEqual(candidate.headers, ("거래처명", "결제금액", "거래일"))
            result = load_detected_table(path, candidate)
            self.assertEqual(result.shape, (2, 3))
            self.assertEqual(result.iloc[1]["결제금액"], 200000)


if __name__ == "__main__":
    unittest.main()
