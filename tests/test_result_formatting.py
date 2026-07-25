from datetime import datetime
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_assistant.excel_io import save_result
from excel_assistant.models import TableCandidate


class ResultFormattingTests(unittest.TestCase):
    def test_source_number_formats_are_applied_to_result_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "source.xlsx"
            result_path = Path(directory) / "result.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "판매"
            ws.append(["날짜", "금액", "비율"])
            ws.append([datetime(2026, 1, 2), 1234567, 0.125])
            ws["A2"].number_format = "mm-dd-yy"
            ws["B2"].number_format = "#,##0"
            ws["C2"].number_format = "0.00%"
            workbook.save(source_path)

            table = TableCandidate(
                sheet_name="판매",
                header_row=1,
                start_column=1,
                end_column=3,
                data_start_row=2,
                data_end_row=2,
                headers=("날짜", "금액", "비율"),
                nonempty_row_count=1,
                confidence=1.0,
            )
            result = pd.DataFrame(
                {
                    "날짜": [pd.Timestamp("2026-01-02")],
                    "금액": [1234567],
                    "비율": [0.125],
                    "새날짜": [pd.Timestamp("2026-02-03")],
                }
            )
            source_bytes_before = source_path.read_bytes()

            save_result(
                result,
                result_path,
                source_path=source_path,
                source_table=table,
            )

            saved = load_workbook(result_path)
            try:
                self.assertEqual(saved.sheetnames, ["결과"])
                result_sheet = saved["결과"]
                self.assertEqual(result_sheet["A2"].number_format, "mm-dd-yy")
                self.assertEqual(result_sheet["B2"].number_format, "#,##0")
                self.assertEqual(result_sheet["C2"].number_format, "0.00%")
                self.assertEqual(result_sheet["D2"].number_format, "yyyy-mm-dd")
            finally:
                saved.close()

            original = load_workbook(source_path)
            try:
                self.assertEqual(original["판매"]["A2"].number_format, "mm-dd-yy")
            finally:
                original.close()
            self.assertEqual(source_path.read_bytes(), source_bytes_before)

    def test_refuses_to_save_over_source_file(self):
        with tempfile.TemporaryDirectory() as directory:
            source_path = Path(directory) / "source.xlsx"
            workbook = Workbook()
            workbook.save(source_path)
            source_bytes_before = source_path.read_bytes()

            with self.assertRaises(ValueError):
                save_result(
                    pd.DataFrame({"value": [1]}),
                    source_path,
                    source_path=source_path,
                )

            self.assertEqual(source_path.read_bytes(), source_bytes_before)


if __name__ == "__main__":
    unittest.main()
